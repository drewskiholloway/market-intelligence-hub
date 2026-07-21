#!/usr/bin/env python3
"""
Fetches the latest Census Bureau New Residential Construction (NRC) release
workbook and extracts the trailing ~13 months of national, seasonally-adjusted
annual-rate (SAAR) Starts and Permits totals into data/permits/national.json.

Source: https://www.census.gov/construction/nrc/current/index.html
        https://www.census.gov/construction/nrc/xls/newresconst.xlsx

This release workbook only ever contains a rolling ~13-month trailing window
(not full history), so the dashboard treats this as an overlay on top of its
long-run baked-in starts/permits series, not a replacement.
"""
import datetime
import json
import re
import subprocess
import sys
import tempfile
import time
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import refresh_permits_data as rpd  # reuse UA / log()

DATA_DIR = rpd.REPO_ROOT / "data" / "permits"
NRC_URL = "https://www.census.gov/construction/nrc/xls/newresconst.xlsx"

MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June",
     "July", "August", "September", "October", "November", "December"], start=1)}


def download_xlsx(url, retries=4):
    last_err = None
    for attempt in range(retries):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = tmp.name
        try:
            subprocess.run(
                ["curl", "-sfL", "--retry", "2", "--max-time", "90",
                 "-A", rpd.UA, "-o", tmp_path, url],
                check=True,
            )
            data = Path(tmp_path).read_bytes()
            if len(data) < 1000:
                raise ValueError(f"Suspiciously small download ({len(data)} bytes) from {url}")
            return data
        except Exception as e:
            last_err = e
            time.sleep(2 * (attempt + 1))
        finally:
            Path(tmp_path).unlink(missing_ok=True)
    raise last_err


def parse_sa_total(wb, sheet_name):
    sh = wb[sheet_name]
    out = {}
    year = None
    started = False
    for r in range(1, sh.max_row + 1):
        label = sh.cell(r, 1).value
        if isinstance(label, str) and re.match(r'^Table \d[a-z] -', label):
            if 'Seasonally adjusted annual rate' in label:
                started = True
                continue
            elif started:
                break
        if not started:
            continue
        if label is None or label == '':
            continue
        if isinstance(label, (int, float)):
            year = int(label)
            continue
        if not isinstance(label, str):
            continue
        stripped = label.strip()
        if re.fullmatch(r'\d{4}', stripped):
            year = int(stripped)
            continue
        if stripped.startswith('.') or stripped.lower().startswith('period'):
            continue
        m = re.match(r'^([A-Za-z]+)', stripped)
        if not m or m.group(1).lower() not in MONTHS or year is None:
            continue
        month = MONTHS[m.group(1).lower()]
        total = sh.cell(r, 2).value
        if not isinstance(total, (int, float)):
            continue
        out[f"{year}-{month:02d}"] = total
    return out


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    rpd.log("refresh_national_data: fetching NRC current release ...")
    raw = download_xlsx(NRC_URL)
    wb = openpyxl.load_workbook(__import__("io").BytesIO(raw), data_only=True)

    starts = parse_sa_total(wb, "Table 3 - Starts")
    permits = parse_sa_total(wb, "Table 1 - Permits")

    payload = {
        "source": "Census Bureau New Residential Construction (NRC)",
        "sourceUrl": NRC_URL,
        "fetchedAt": datetime.datetime.now().isoformat(timespec="seconds"),
        "unit": "SAAR (thousands of units), United States total",
        "data": {"starts": starts, "permits": permits},
    }

    (DATA_DIR / "national.json").write_text(json.dumps(payload, indent=2))
    rpd.log(f"refresh_national_data: stored {len(starts)} months of starts, "
             f"{len(permits)} months of permits (latest: "
             f"{max(starts) if starts else 'n/a'}).")


if __name__ == "__main__":
    main()
